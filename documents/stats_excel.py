"""Professional activity-report (bilan) Excel generator.

Renders the configured bilan (see :mod:`core.bilan`) as a presentation-ready
``.xlsx`` workbook following common institutional-reporting practice:

  * a ``Synthèse`` cover sheet — institution header, reporting period,
    generation metadata, and the headline KPIs (counts + amounts);
  * one sheet per selected dimension, each a self-contained table with a
    styled header, banded rows, a share-of-total (%) column, per-channel and
    grand-total amounts, a bold totals row, frozen header, auto-filter and
    currency/percent number formats.

All monetary values are expressed in Algerian Dinar (DA).
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from django.conf import settings

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference


# Brand palette
_NAVY = '1E293B'
_INDIGO = '4F46E5'
_INDIGO_LT = 'EEF2FF'
_BAND = 'F8FAFC'
_GREY = '64748B'

_MONEY_FMT = '#,##0 "DA"'
_PCT_FMT = '0.0"%"'
_INT_FMT = '#,##0'

_thin = Side(style='thin', color='E2E8F0')
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _title_font(size=16, color=_NAVY):
    return Font(name='Calibri', size=size, bold=True, color=color)


def _hdr_fill():
    return PatternFill('solid', fgColor=_INDIGO)


def _money_cols(columns):
    return [i for i, c in enumerate(columns) if '(DA)' in c]


def _write_section(ws, section, start_row=1):
    cols = section['columns']
    ncols = len(cols)
    money = set(_money_cols(cols))

    # Section title
    ws.cell(row=start_row, column=1, value=section['title']).font = _title_font(13)
    r = start_row + 1

    # Header row
    for c, name in enumerate(cols, start=1):
        cell = ws.cell(row=r, column=c, value=name)
        cell.fill = _hdr_fill()
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _BORDER
    header_row = r
    r += 1

    # Data rows (banded)
    for i, row in enumerate(section['rows']):
        band = PatternFill('solid', fgColor=_BAND) if i % 2 else None
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = _BORDER
            if band:
                cell.fill = band
            idx = c - 1
            if idx == 0:
                cell.alignment = Alignment(horizontal='left')
            elif idx == 2:                       # Part (%)
                cell.number_format = _PCT_FMT
                cell.alignment = Alignment(horizontal='right')
            elif idx in money:
                cell.number_format = _MONEY_FMT
                cell.alignment = Alignment(horizontal='right')
            else:                                 # counts
                cell.number_format = _INT_FMT
                cell.alignment = Alignment(horizontal='right')
        r += 1

    # Totals row
    for c, val in enumerate(section['total_row'], start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor=_INDIGO_LT)
        cell.border = _BORDER
        idx = c - 1
        if idx == 2:
            cell.number_format = _PCT_FMT
            cell.alignment = Alignment(horizontal='right')
        elif idx in money:
            cell.number_format = _MONEY_FMT
            cell.alignment = Alignment(horizontal='right')
        elif idx >= 1:
            cell.number_format = _INT_FMT
            cell.alignment = Alignment(horizontal='right')
    totals_row = r

    # Frozen header + auto-filter + widths
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = (
        f"{get_column_letter(1)}{header_row}:"
        f"{get_column_letter(ncols)}{totals_row - 1}"
    )
    widths = [40, 12, 12] + [22] * (ncols - 3)
    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = widths[c - 1] if c - 1 < len(widths) else 18

    # Figure: horizontal bar chart of the counts by category (top 15),
    # placed to the right of the table.
    first_data = header_row + 1
    last_data = totals_row - 1
    if last_data >= first_data:
        max_row = min(last_data, first_data + 14)
        chart = BarChart()
        chart.type = 'bar'
        chart.style = 10
        chart.title = section['title']
        chart.legend = None
        data = Reference(ws, min_col=2, max_col=2, min_row=header_row, max_row=max_row)
        cats = Reference(ws, min_col=1, min_row=first_data, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = max(6.0, (max_row - first_data + 1) * 0.55)
        chart.width = 15
        ws.add_chart(chart, f"{get_column_letter(ncols + 2)}{header_row}")
    return totals_row


def _safe_sheet_name(name, used):
    bad = set('[]:*?/\\')
    clean = ''.join(ch for ch in name if ch not in bad)[:28].strip() or 'Section'
    base, n = clean, 2
    while clean in used:
        clean = f"{base[:25]}-{n}"
        n += 1
    used.add(clean)
    return clean


def _period_label(filters: dict) -> str:
    df = filters.get('date_from')
    dt = filters.get('date_to')
    if df and dt:
        return f"du {df} au {dt}"
    if df:
        return f"à partir du {df}"
    if dt:
        return f"jusqu'au {dt}"
    return "toutes périodes confondues"


def generate_bilan_excel(bilan: dict, filters: dict, actor) -> str:
    """Render ``bilan`` to an .xlsx workbook and return its path."""
    wb = Workbook()

    # ---- Synthèse sheet -------------------------------------------------
    ws = wb.active
    ws.title = 'Synthèse'
    ws.sheet_view.showGridLines = False

    ws.cell(row=1, column=1, value="ESSBO — École Supérieure en Sciences Biologiques d'Oran").font = _title_font(12, _GREY)
    ws.cell(row=2, column=1, value="PLAGENOR 4.0 — Bilan d'activité").font = _title_font(18)
    gran_lbl = {'month': 'mensuelle', 'quarter': 'trimestrielle',
                'year': 'annuelle'}.get(bilan.get('granularity'), 'mensuelle')
    ws.cell(row=3, column=1,
            value=f"Période : {_period_label(filters)} — granularité {gran_lbl}").font = Font(size=11, color=_GREY)
    ws.cell(row=4, column=1,
            value=(f"Édité le {datetime.now().strftime('%d/%m/%Y à %H:%M')} "
                   f"par {actor.get_full_name() or actor.username}")).font = Font(size=10, color=_GREY)

    k = bilan['kpis']
    kpi_rows = [
        ("Total des demandes", k['total'], _INT_FMT),
        ("Demandes complétées", k['completed'], _INT_FMT),
        ("En cours", k['in_progress'], _INT_FMT),
        ("Rejetées / brouillons", k['rejected'], _INT_FMT),
        ("Taux de complétion", k['completion_rate'], _PCT_FMT),
        ("Demandes IBTIKAR", k['ibtikar_count'], _INT_FMT),
        ("Demandes GENOCLAB", k['genoclab_count'], _INT_FMT),
        ("Valeur virtuelle IBTIKAR", k['ibtikar_virtual_revenue'], _MONEY_FMT),
        ("Chiffre d'affaires GENOCLAB", k['genoclab_revenue'], _MONEY_FMT),
    ]
    r = 6
    ws.cell(row=r, column=1, value="Indicateurs clés").font = _title_font(13)
    r += 1
    ib_row = gc_row = None
    for label, value, fmt in kpi_rows:
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(bold=True)
        lc.fill = PatternFill('solid', fgColor=_INDIGO_LT)
        lc.border = _BORDER
        vc = ws.cell(row=r, column=2, value=value)
        vc.number_format = fmt
        vc.alignment = Alignment(horizontal='right')
        vc.border = _BORDER
        if label == "Demandes IBTIKAR":
            ib_row = r
        elif label == "Demandes GENOCLAB":
            gc_row = r
        r += 1
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 22

    # Figure: répartition des demandes par canal (IBTIKAR vs GENOCLAB).
    if ib_row and gc_row == ib_row + 1 and (k['ibtikar_count'] or k['genoclab_count']):
        pie = PieChart()
        pie.title = "Répartition par canal"
        data = Reference(ws, min_col=2, min_row=ib_row, max_row=gc_row)
        cats = Reference(ws, min_col=1, min_row=ib_row, max_row=gc_row)
        pie.add_data(data, titles_from_data=False)
        pie.set_categories(cats)
        pie.height = 7
        pie.width = 11
        ws.add_chart(pie, "D6")

    # ---- One sheet per section -----------------------------------------
    used = {'Synthèse'}
    for section in bilan['sections']:
        sheet = wb.create_sheet(_safe_sheet_name(section['title'], used))
        sheet.sheet_view.showGridLines = False
        _write_section(sheet, section, start_row=1)

    # ---- Persist --------------------------------------------------------
    out_dir = Path(settings.MEDIA_ROOT) / 'documents'
    out_dir.mkdir(parents=True, exist_ok=True)
    fname = f"PLAGENOR_Bilan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = out_dir / fname
    wb.save(path)
    return str(path)
