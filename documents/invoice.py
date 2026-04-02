# documents/invoice.py — PLAGENOR 4.0 Invoice PDF/Excel Generator
from io import BytesIO
from datetime import datetime
from django.conf import settings


def generate_invoice_pdf(invoice):
    """
    Generate a simple PDF invoice for a GENOCLAB request using ReportLab.
    Args:
        invoice: Invoice model instance
    Returns:
        bytes: PDF file content
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm, mm
    from reportlab.lib.colors import HexColor, black
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=2*cm,
        rightMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle('Title2', parent=styles['Heading1'], fontSize=18, spaceAfter=12, alignment=TA_CENTER))
    styles.add(ParagraphStyle('SubTitle', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, textColor=HexColor('#666666')))
    styles.add(ParagraphStyle('InvoiceLabel', parent=styles['Normal'], fontSize=9, textColor=HexColor('#888888')))

    story = []

    # Header
    story.append(Paragraph("FACTURE / INVOICE", styles['Title2']))
    story.append(Paragraph("PLAGENOR 4.0 — Plateforme Génomique", styles['SubTitle']))
    story.append(Spacer(1, 8))

    # Invoice meta
    meta_data = [
        ["N° Facture / Invoice No:", invoice.invoice_number],
        ["Date:", invoice.created_at.strftime('%d/%m/%Y') if invoice.created_at else datetime.now().strftime('%d/%m/%Y')],
        ["Client:", invoice.client.get_full_name() if invoice.client else str(invoice.client)],
        ["Demande / Request:", invoice.request.display_id if invoice.request else ''],
    ]
    meta_table = Table(meta_data, colWidths=[5*cm, 12*cm])
    meta_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 16))

    # Line items table
    story.append(Paragraph("Détails / Details", styles['Normal']))
    story.append(Spacer(1, 6))

    items = invoice.line_items or []
    if not items and invoice.request:
        items = [{
            'description': invoice.request.title,
            'quantity': 1,
            'unit_price': float(invoice.subtotal_ht or 0),
        }]

    table_data = [
        [Paragraph('<b>Description</b>', styles['Normal']),
         Paragraph('<b>Qté</b>', styles['Normal']),
         Paragraph('<b>Prix Unitaire</b>', styles['Normal']),
         Paragraph('<b>Total HT</b>', styles['Normal'])],
    ]
    for item in items:
        desc = item.get('description', '')
        qty = item.get('quantity', 1)
        unit = item.get('unit_price', 0)
        total = qty * unit
        table_data.append([
            Paragraph(str(desc), styles['Normal']),
            Paragraph(str(qty), styles['Normal']),
            Paragraph(f"{float(unit):,.2f} DA", styles['Normal']),
            Paragraph(f"{float(total):,.2f} DA", styles['Normal']),
        ])

    item_table = Table(table_data, colWidths=[8*cm, 2*cm, 4*cm, 4*cm])
    item_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#f0f0f0')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#cccccc')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(item_table)
    story.append(Spacer(1, 12))

    # Totals
    totals = [
        [Paragraph('Sous-total HT / Subtotal:', styles['Normal']),
         Paragraph(f"{float(invoice.subtotal_ht or 0):,.2f} DA", styles['Normal'])],
        [Paragraph(f"TVA ({float(invoice.vat_rate or 0)*100:.0f}%) / VAT:", styles['Normal']),
         Paragraph(f"{float(invoice.vat_amount or 0):,.2f} DA", styles['Normal'])],
        [Paragraph('<b>Total TTC / Total:</b>', styles['Normal']),
         Paragraph(f"<b>{float(invoice.total_ttc or 0):,.2f} DA</b>", styles['Normal'])],
    ]
    totals_table = Table(totals, colWidths=[12*cm, 6*cm])
    totals_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LINEABOVE', (0, 2), (-1, 2), 1, black),
    ]))
    story.append(totals_table)
    story.append(Spacer(1, 20))

    # Payment info
    try:
        ps = getattr(settings, 'PAYMENT_SETTINGS', {})
        if ps.get('bank_account'):
            story.append(Paragraph("Informations de paiement / Payment Information", styles['Normal']))
            story.append(Spacer(1, 6))
            story.append(Paragraph(f"Banque: {ps.get('bank_name', '')}", styles['Normal']))
            story.append(Paragraph(f"Compte: {ps.get('bank_account', '')}", styles['Normal']))
            story.append(Paragraph(f"Bénéficiaire: {ps.get('beneficiary_name', '')}", styles['Normal']))
    except Exception:
        pass

    doc.build(story)
    return buffer.getvalue()


def generate_invoice_excel(request_obj):
    """
    Generate a simple Excel invoice for a GENOCLAB request using openpyxl.
    Args:
        request_obj: Request model instance
    Returns:
        bytes: Excel file content
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Facture"

    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Title
    ws['A1'] = "FACTURE — INVOICE"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')

    # Invoice details
    row = 3
    ws[f'A{row}'] = "N° Facture:"
    ws[f'B{row}'] = f"INV-{request_obj.display_id}-{datetime.now().strftime('%Y%m%d')}"
    ws[f'A{row}'].font = bold_font
    row += 1
    ws[f'A{row}'] = "Date:"
    ws[f'B{row}'] = datetime.now().strftime('%d/%m/%Y')
    ws[f'A{row}'].font = bold_font
    row += 1
    ws[f'A{row}'] = "Client:"
    ws[f'B{row}'] = request_obj.requester.get_full_name() if request_obj.requester else str(request_obj.requester)
    ws[f'A{row}'].font = bold_font
    row += 1
    ws[f'A{row}'] = "Demande:"
    ws[f'B{row}'] = request_obj.display_id
    ws[f'A{row}'].font = bold_font
    row += 2

    # Table headers
    headers = ["Description", "Qté / Quantity", "Prix Unitaire / Unit Price", "Total HT"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    row += 1

    # Line items
    items = []
    if request_obj.quote_detail and request_obj.quote_detail.get('items'):
        items = request_obj.quote_detail.get('items', [])
    elif request_obj.quote_amount:
        items = [{'description': request_obj.title, 'quantity': 1, 'unit_price': float(request_obj.quote_amount)}]

    for item in items:
        desc = item.get('description', '')
        qty = item.get('quantity', 1)
        unit = item.get('unit_price', 0)
        total = qty * unit
        for col, val in enumerate([desc, qty, f"{unit:,.2f} DA", f"{total:,.2f} DA"], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            if col > 1:
                cell.alignment = Alignment(horizontal='right')
        row += 1

    # Totals
    row += 1
    subtotal = sum(i.get('quantity', 1) * i.get('unit_price', 0) for i in items)
    vat = round(subtotal * float(getattr(settings, 'VAT_RATE', 0.19)), 2)
    total = subtotal + vat

    for label, val in [
        ("Sous-total HT / Subtotal:", f"{subtotal:,.2f} DA"),
        (f"TVA / VAT ({float(getattr(settings, 'VAT_RATE', 0.19))*100:.0f}%):", f"{vat:,.2f} DA"),
        ("Total TTC / Total:", f"{total:,.2f} DA"),
    ]:
        ws[f'A{row}'] = label
        ws[f'D{row}'] = val
        ws[f'A{row}'].font = bold_font
        ws[f'D{row}'].alignment = Alignment(horizontal='right')
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
